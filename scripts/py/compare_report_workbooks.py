"""Compare two Excel report workbooks for regression testing.

The script is for development on the external-network machine. It is not part of
the VBA runtime and never needs to be copied into the internal network.
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class Difference:
    location: str
    expected: Any
    actual: Any


def normalized(value: Any) -> Any:
    if isinstance(value, float) and math.isnan(value):
        return "<NaN>"
    return value


def compare_workbooks(expected_path: Path, actual_path: Path) -> list[Difference]:
    expected = load_workbook(expected_path, data_only=False, read_only=False)
    actual = load_workbook(actual_path, data_only=False, read_only=False)
    differences: list[Difference] = []

    try:
        if expected.sheetnames != actual.sheetnames:
            differences.append(
                Difference("workbook.sheetnames", expected.sheetnames, actual.sheetnames)
            )

        for sheet_name in expected.sheetnames:
            if sheet_name not in actual.sheetnames:
                continue
            expected_sheet = expected[sheet_name]
            actual_sheet = actual[sheet_name]

            if expected_sheet.max_row != actual_sheet.max_row:
                differences.append(
                    Difference(
                        f"{sheet_name}.max_row",
                        expected_sheet.max_row,
                        actual_sheet.max_row,
                    )
                )
            if expected_sheet.max_column != actual_sheet.max_column:
                differences.append(
                    Difference(
                        f"{sheet_name}.max_column",
                        expected_sheet.max_column,
                        actual_sheet.max_column,
                    )
                )

            expected_merges = sorted(str(item) for item in expected_sheet.merged_cells.ranges)
            actual_merges = sorted(str(item) for item in actual_sheet.merged_cells.ranges)
            if expected_merges != actual_merges:
                differences.append(
                    Difference(f"{sheet_name}.merged_cells", expected_merges, actual_merges)
                )

            max_row = max(expected_sheet.max_row, actual_sheet.max_row)
            max_column = max(expected_sheet.max_column, actual_sheet.max_column)
            for row in range(1, max_row + 1):
                for column in range(1, max_column + 1):
                    expected_value = normalized(expected_sheet.cell(row, column).value)
                    actual_value = normalized(actual_sheet.cell(row, column).value)
                    if expected_value != actual_value:
                        coordinate = expected_sheet.cell(row, column).coordinate
                        differences.append(
                            Difference(
                                f"{sheet_name}!{coordinate}", expected_value, actual_value
                            )
                        )
    finally:
        expected.close()
        actual.close()

    return differences


def print_differences(differences: Iterable[Difference], maximum: int) -> int:
    all_differences = list(differences)
    for difference in all_differences[:maximum]:
        print(f"{difference.location}")
        print(f"  expected: {difference.expected!r}")
        print(f"  actual:   {difference.actual!r}")
    if len(all_differences) > maximum:
        print(f"... {len(all_differences) - maximum} additional differences omitted")
    return len(all_differences)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("expected", type=Path)
    parser.add_argument("actual", type=Path)
    parser.add_argument("--max-differences", type=int, default=200)
    args = parser.parse_args()

    differences = compare_workbooks(args.expected, args.actual)
    count = print_differences(differences, args.max_differences)
    if count:
        print(f"FAILED: {count} difference(s)")
        return 1
    print("OK: workbooks match")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
